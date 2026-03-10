#!/usr/bin/env python3
"""
Script to read and process Excel .xlsm files

This utility script provides functions to inspect and read Excel files,
particularly useful for understanding the structure of data files before
processing them in the seed script.
"""

import sys
from pathlib import Path
from typing import Dict, Optional

try:
    import openpyxl
    import pandas as pd
except ImportError as e:
    print(f"Error: Missing required dependency: {e}", file=sys.stderr)
    print("Please install dependencies: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(1)


def get_excel_info(file_path: Path) -> openpyxl.Workbook:
    """
    Get basic information about the Excel file
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        openpyxl.Workbook object
        
    Raises:
        FileNotFoundError: If file doesn't exist
        openpyxl.utils.exceptions.InvalidFileException: If file is invalid
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    print(f"Reading Excel file: {file_path}\n")
    
    try:
        # Load workbook with data_only=True to get calculated values
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}", file=sys.stderr)
        raise
    
    print(f"Number of sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}\n")
    
    return wb


def read_sheet_as_dataframe(file_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Read a specific sheet as a pandas DataFrame
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (None for first sheet)
        
    Returns:
        pandas DataFrame with sheet data
        
    Raises:
        ValueError: If sheet_name doesn't exist
    """
    try:
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        else:
            # Read first sheet by default
            df = pd.read_excel(file_path, engine='openpyxl')
        return df
    except ValueError as e:
        print(f"Error reading sheet '{sheet_name}': {e}", file=sys.stderr)
        raise


def display_sheet_info(file_path, sheet_name):
    """Display information about a specific sheet"""
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    df = read_sheet_as_dataframe(file_path, sheet_name)
    
    print(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
    print(f"\nColumn names:")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i}. {col}")
    
    print(f"\nFirst few rows:")
    print(df.head(10).to_string())
    
    print(f"\nData types:")
    print(df.dtypes)
    
    print(f"\nSummary statistics:")
    print(df.describe())
    
    return df


def list_all_sheets(file_path):
    """List all sheets and their basic info"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print("All sheets in the workbook:\n")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  - {sheet_name}: {ws.max_row} rows × {ws.max_column} columns")
    
    return wb.sheetnames


def read_all_sheets(file_path):
    """Read all sheets and return as a dictionary of DataFrames"""
    excel_file = pd.ExcelFile(file_path, engine='openpyxl')
    sheets_dict = {}
    
    for sheet_name in excel_file.sheet_names:
        sheets_dict[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    return sheets_dict


def main() -> Dict[str, pd.DataFrame]:
    """
    Main function to process Excel file and display information
    
    Returns:
        Dictionary mapping sheet names to DataFrames
        
    Raises:
        SystemExit: If file not found or processing fails
    """
    # Excel file path (relative to project root)
    excel_file = Path("data/galaticos.xlsm")
    
    # Try to find file in current directory or data/raw
    if not excel_file.exists():
        alt_path = Path("data") / excel_file.name
        if alt_path.exists():
            excel_file = alt_path
        else:
            print(f"Error: File '{excel_file}' not found!", file=sys.stderr)
            print(f"Also checked: {alt_path}", file=sys.stderr)
            sys.exit(1)
    
    try:
        # Get basic info
        wb = get_excel_info(excel_file)
        
        # List all sheets
        print("\n" + "="*60)
        print("SHEET INFORMATION")
        print("="*60)
        sheet_names = list_all_sheets(excel_file)
        
        # Display detailed info for each sheet
        print("\n" + "="*60)
        print("DETAILED SHEET INFORMATION")
        print("="*60)
        
        all_sheets = read_all_sheets(excel_file)
        
        for sheet_name in sheet_names:
            display_sheet_info(excel_file, sheet_name)
        
        # Return the sheets dictionary for further processing
        return all_sheets
    except Exception as e:
        print(f"Error processing Excel file: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    sheets_data = main()
    
    # Example: Access specific sheet data
    # if 'Sheet1' in sheets_data:
    #     df = sheets_data['Sheet1']
    #     # Process your data here
    #     print("\nProcessing complete!")

